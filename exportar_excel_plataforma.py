"""
Exporta a Excel todos los objetos de plataforma.sistemagps.online
con: Grupo, Placa, IMEI, Último Reporte
"""

import requests
import json
from datetime import datetime
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ---------- CONFIGURACIÓN ----------
BASE       = 'https://plataforma.sistemagps.online/api'
EMAIL      = 'gerencia@rastrear.com.co'
PASSWORD   = 'a791025*'
EXCEL_PATH = 'reporte_plataforma_v2.xlsx'
# -----------------------------------

def make_session():
    s = requests.Session()
    s.mount('https://', HTTPAdapter(max_retries=Retry(total=3, backoff_factor=1)))
    return s

def login(session):
    r = session.post(f'{BASE}/login', json={'email': EMAIL, 'password': PASSWORD}, timeout=30)
    r.raise_for_status()
    data = r.json()
    if data.get('status') != 1:
        raise ValueError(f'Login fallido: {data}')
    return data['user_api_hash']

def get_devices(session, token):
    r = session.post(f'{BASE}/get_devices', json={'user_api_hash': token}, timeout=60)
    r.raise_for_status()
    return r.json()

def parse_last_report(device):
    """Intenta extraer la fecha/hora del último reporte del dispositivo."""
    for campo in ['time', 'last_connect', 'dt_tracker', 'lat_time', 'server_time', 'dt_server']:
        val = device.get(campo)
        if val:
            return str(val)
    dd = device.get('device_data') or {}
    for campo in ['time', 'last_connect', 'dt_tracker', 'server_time']:
        val = dd.get(campo)
        if val:
            return str(val)
    return 'Sin datos'

def parse_creation_date(device):
    """Intenta extraer la fecha de creación del dispositivo."""
    for campo in ['created_at', 'created', 'dt_created', 'registration_date', 'date_added', 'added_at']:
        val = device.get(campo)
        if val:
            return str(val)
    dd = device.get('device_data') or {}
    for campo in ['created_at', 'created', 'dt_created', 'registration_date']:
        val = dd.get(campo)
        if val:
            return str(val)
    return 'Sin datos'

def main():
    try:
        import openpyxl
    except ImportError:
        print('Instalando openpyxl...')
        import subprocess, sys
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
        import openpyxl

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    session = make_session()

    print('🔐 Iniciando sesión...')
    token = login(session)
    print(f'   Token: {token[:12]}...')

    print('📡 Obteniendo dispositivos...')
    grupos = get_devices(session, token)
    print(f'   Grupos encontrados: {len(grupos)}')

    # Recopilar filas
    filas = []
    for grupo in grupos:
        grupo_nombre = grupo.get('title') or grupo.get('name') or f'Grupo {grupo.get("id","?")}'
        items = grupo.get('items', [])
        for dev in items:
            placa        = dev.get('name', '').strip()
            imei         = (dev.get('device_data') or {}).get('imei', '') or dev.get('imei', '')
            ultimo_rep   = parse_last_report(dev)
            fecha_crea   = parse_creation_date(dev)
            filas.append({
                'Grupo'            : grupo_nombre,
                'Placa'            : placa,
                'IMEI'             : imei,
                'Último Reporte'   : ultimo_rep,
                'Fecha de Creación': fecha_crea,
            })

    print(f'   Total de objetos: {len(filas)}')

    if not filas:
        print('⚠  No se encontraron dispositivos. Verifica las credenciales.')
        return

    # -------- Crear Excel --------
    wb = Workbook()
    ws = wb.active
    ws.title = 'Plataforma GPS'

    # Estilos
    header_fill   = PatternFill('solid', fgColor='1F4E79')
    header_font   = Font(bold=True, color='FFFFFF', size=11)
    header_align  = Alignment(horizontal='center', vertical='center')
    cell_align    = Alignment(horizontal='left', vertical='center', wrap_text=False)
    thin_border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alt_fill = PatternFill('solid', fgColor='D6E4F0')

    columnas = ['Grupo', 'Placa', 'IMEI', 'Último Reporte', 'Fecha de Creación']

    # Encabezado
    ws.row_dimensions[1].height = 22
    for col_idx, nombre_col in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=nombre_col)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = thin_border

    # Datos
    for row_idx, fila in enumerate(filas, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        ws.row_dimensions[row_idx].height = 16
        for col_idx, col_key in enumerate(columnas, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=fila[col_key])
            cell.alignment = cell_align
            cell.border    = thin_border
            if fill:
                cell.fill = fill

    # Ancho automático de columnas
    anchos = {'Grupo': 30, 'Placa': 18, 'IMEI': 22, 'Último Reporte': 25, 'Fecha de Creación': 25}
    for col_idx, col_key in enumerate(columnas, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = anchos[col_key]

    # Título general
    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.insert_rows(1)
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value     = f'Reporte de Objetos — plataforma.sistemagps.online  ({now_str})'
    title_cell.font      = Font(bold=True, size=13, color='1F4E79')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Congelar encabezado
    ws.freeze_panes = 'A3'

    # Guardar
    wb.save(EXCEL_PATH)
    print(f'\n✅ Excel generado: {EXCEL_PATH}  ({len(filas)} objetos)')

if __name__ == '__main__':
    main()
